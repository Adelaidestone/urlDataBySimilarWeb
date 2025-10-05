import json
import openpyxl
from datetime import datetime
from urllib.parse import urlparse

# 文件路径
txt_file_path = r'E:\aiUrlDataBySimilarWeb\similarweb_data.txt'
excel_file_path = r'D:\Users\Mussy\Desktop\similarweb_data.xlsx'

def extract_domain(url):
    """
    从URL中提取主域名，去除协议、www、路径、参数等
    例如: https://www.example.com/path?param=value -> example.com
    """
    url = url.strip().lower()
    
    # 添加协议如果没有的话（用于urlparse解析）
    if not url.startswith(('http://', 'https://')):
        url = 'http://' + url
    
    try:
        parsed = urlparse(url)
        domain = parsed.netloc if parsed.netloc else parsed.path.split('/')[0]
        
        # 移除www前缀
        if domain.startswith('www.'):
            domain = domain[4:]
        
        return domain
    except:
        # 如果解析失败，手动处理
        url = url.replace('https://', '').replace('http://', '')
        domain = url.split('/')[0].split('?')[0]
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain

def domains_match(domain1, domain2):
    """
    判断两个域名是否完全匹配
    只有完全相等才返回True
    """
    return domain1 == domain2

print("正在读取JSON数据...")
# 1. 读取similarweb_data.txt中的JSON数据
json_data_dict = {}
with open(txt_file_path, 'r', encoding='utf-8') as f:
    for line in f:
        if line.strip():
            data = json.loads(line)
            url = list(data.keys())[0]
            json_data_dict[url] = data[url]

print(f"已读取 {len(json_data_dict)} 条JSON数据")

# 2. 加载Excel文件
print("\n正在加载Excel文件...")
wb = openpyxl.load_workbook(excel_file_path)
ws = wb.active

# 列索引映射（根据之前的输出）
COL_URL = 2  # 官网链接
COL_DESKTOP = 3  # 桌面端占比
COL_MOBILE = 4  # 移动端占比
COL_VISITS = 5  # 每月访问量
COL_VISITS_PER_VISITOR = 6  # 每访客访问次数
COL_USERS_TAB = 7  # 已消除重叠的受众
COL_PAGES_PER_VISIT = 8  # 页面数/访问
COL_AVG_VISIT_DURATION = 9  # 访问持续时间
COL_BOUNCE_RATE = 10  # 跳出率
COL_UPDATE_TIME = 11  # 更新时间

# 3. 遍历Excel的每一行，查找匹配的URL
print("\n开始匹配和填充数据...")
matched_count = 0
not_matched_count = 0
current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

for row in range(2, ws.max_row + 1):  # 从第2行开始（第1行是标题）
    excel_url = ws.cell(row, COL_URL).value
    
    if not excel_url:
        continue
    
    # 提取Excel URL的域名
    excel_domain = extract_domain(excel_url)
    
    # 在JSON数据中查找匹配的URL
    matched_data = None
    matched_json_url = None
    for json_url, json_values in json_data_dict.items():
        json_domain = extract_domain(json_url)
        
        # 使用精确的域名匹配
        if domains_match(excel_domain, json_domain):
            matched_data = json_values
            matched_json_url = json_url
            break
    
    if matched_data:
        # 填充数据
        ws.cell(row, COL_DESKTOP).value = matched_data.get('desktopPersent', 'N/A')
        ws.cell(row, COL_MOBILE).value = matched_data.get('mobilePercent', 'N/A')
        ws.cell(row, COL_VISITS).value = matched_data.get('visits', 0)
        ws.cell(row, COL_VISITS_PER_VISITOR).value = matched_data.get('visits_per_visitor', 0)
        ws.cell(row, COL_USERS_TAB).value = matched_data.get('users_tab', 0)
        ws.cell(row, COL_PAGES_PER_VISIT).value = matched_data.get('pages-per-visit', 0)
        ws.cell(row, COL_AVG_VISIT_DURATION).value = matched_data.get('avg_visit_duration', 'N/A')
        ws.cell(row, COL_BOUNCE_RATE).value = matched_data.get('bounce_rate', 'N/A')
        ws.cell(row, COL_UPDATE_TIME).value = current_time
        
        matched_count += 1
        product_name = ws.cell(row, 1).value
        # 显示域名匹配信息（Excel域名 ← JSON域名）
        if excel_domain != extract_domain(matched_json_url):
            print(f"✓ 行{row}: {product_name} ({excel_domain} ← {matched_json_url}) - 已更新")
        else:
            print(f"✓ 行{row}: {product_name} ({excel_domain}) - 已更新")
    else:
        not_matched_count += 1
        # print(f"✗ 行{row}: {excel_domain} - 未找到匹配数据")

# 4. 保存Excel文件
print(f"\n正在保存Excel文件...")
import time
max_retries = 3
for attempt in range(max_retries):
    try:
        wb.save(excel_file_path)
        print("✓ Excel文件保存成功！")
        break
    except PermissionError:
        if attempt < max_retries - 1:
            print(f"\n⚠️  Excel文件正在被使用，无法保存。")
            print(f"请关闭 similarweb_data.xlsx 文件，然后按回车继续...")
            input()
            print(f"正在重试保存... (第 {attempt + 2} 次尝试)")
        else:
            print("\n" + "!"*60)
            print("❌ 错误：Excel文件一直处于打开状态，无法保存！")
            print("请关闭 similarweb_data.xlsx 文件后重新运行此脚本。")
            print("!"*60)
            exit(1)
    except Exception as e:
        print(f"❌ 保存失败: {e}")
        exit(1)

# 5. 输出统计信息
print("\n" + "="*60)
print("✅ 导入完成！")
print(f"成功匹配并更新: {matched_count} 条")
print(f"未找到匹配: {not_matched_count} 条")
print(f"总行数: {ws.max_row - 1} 行（不含标题）")
print(f"更新时间: {current_time}")
print("="*60)

